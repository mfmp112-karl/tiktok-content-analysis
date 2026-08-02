"""The calendar as a spreadsheet you can actually work in.

Tabs follow the source template's structure — how to use it, what each post type
is for, the thirty-day grid, a hook library and a progress tracker — because
that structure is the method and it works. The grid's left side is filled in
from the analysis; the right side is left empty for the creator's own caption,
platform and a tick when it is posted.

Method credit: themarketingfmpodcast. Every sheet carries the attribution line,
and `attribution.assert_stamped` is run over the workbook's own text before it
is written, so removing the stamping fails the run rather than shipping quietly.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .. import attribution
from ..stats import plain
from . import cycle

INK, INK_SOFT = "0B0B0B", "52514E"
HEADER_BG, BAND_BG, ACCENT_BG = "1F2933", "EEF3F9", "FFF4D6"
GOOD, WARN, MUTED = "0CA30C", "B07800", "7A7973"

TYPE_FILL = {
    "EDUCATE": "E7EEFA", "PROOF": "E4F5E4", "ENGAGE": "FCE9F0",
    "OFFER": "FDF3DC", "MIRROR": "EDEBF7",
}

TITLE_FONT = Font(bold=True, size=16, color=INK)
H2_FONT = Font(bold=True, size=11, color="FFFFFF")
MUTED_FONT = Font(color=INK_SOFT, italic=True, size=9)
BOLD = Font(bold=True, color=INK)

VERDICT_COLOR = {"strong": GOOD, "moderate": WARN}


def _title(ws, text: str, sub: str = "") -> int:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if sub:
        ws["A2"] = sub
        ws["A2"].font = MUTED_FONT
        return 4
    return 3


def _header(ws, row: int, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor=HEADER_BG)
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row, i, text)
        c.font = H2_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26


def _widths(ws, spec: dict[int, int]) -> None:
    for col, width in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _stamp(ws, row: int) -> None:
    c = ws.cell(row, 1, attribution.TAG)
    c.font = MUTED_FONT


def _sheet(wb: Workbook, title: str):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    return ws


# ----------------------------------------------------------------------- tabs

def _how_to_use(wb: Workbook, cal: dict, handle: str) -> None:
    ws = _sheet(wb, "How to use")
    row = _title(ws, "Your 30-day content calendar",
                 f"Built from @{handle}'s own posting history.")
    steps = [
        ("1", "Open the '30-day calendar' tab. The post type and theme for every "
              "day are already filled in."),
        ("2", "Read the prompt, then write your actual caption in the "
              "'Your caption / hook' column. That column is the only one you "
              "have to fill in."),
        ("3", "The 'Suggested hook' column holds openings that already worked on "
              "this account. Reuse the shape, not the exact words."),
        ("4", "Mark the platform and tick 'Posted' when it goes out."),
        ("5", "Use the 'Notes' column while it is fresh — what you would change, "
              "what you noticed. That column is where next month's calendar "
              "comes from."),
        ("6", "Check 'Post type guide' when you are unsure what a type is for, "
              "and 'Hook library' when you are stuck on an opening."),
        ("7", "At the end of the month, look at 'Progress tracker' before you "
              "judge anything. Consistency first, performance second."),
    ]
    _header(ws, row, ["Step", "What to do"])
    row += 1
    for num, text in steps:
        ws.cell(row, 1, num).font = BOLD
        c = ws.cell(row, 2, text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    ws.cell(row, 1, "Why the types rotate").font = BOLD
    row += 1
    c = ws.cell(row, 1, "Left to instinct, almost everyone posts one type over and "
                        "over. The five-day rotation is the whole intervention: it "
                        "forces a spread across teaching, proof, conversation, "
                        "selling and empathy. Six of each, every month.")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 58
    row += 2

    if cal.get("anchor_note"):
        c = ws.cell(row, 1, cal["anchor_note"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 46
        row += 2

    _stamp(ws, row + 1)
    _widths(ws, {1: 10, 2: 92})


def _post_type_guide(wb: Workbook) -> None:
    ws = _sheet(wb, "Post type guide")
    row = _title(ws, "What each post type is for",
                 "Five jobs. Each one reaches a different person.")
    _header(ws, row, ["Type", "What it does", "How to write it"])
    row += 1
    for t in cycle.POST_TYPES:
        c = ws.cell(row, 1, t["name"])
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor=TYPE_FILL[t["name"]])
        c.alignment = Alignment(vertical="top")
        for col, text in ((2, t["does"]), (3, t["how"])):
            cell = ws.cell(row, col, text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 46
        row += 1
    _stamp(ws, row + 1)
    _widths(ws, {1: 14, 2: 52, 3: 52})


def _calendar_tab(wb: Workbook, cal: dict) -> None:
    ws = _sheet(wb, "30-day calendar")
    row = _title(ws, "The next 30 days",
                 "Columns A-I are filled in for you. Columns J-M are yours.")
    headers = ["Day", "Date", "Wk", "Post type", "Theme", "Prompt",
               "Suggested hook", "Best time", "Confidence",
               "Your caption / hook", "Platform", "Posted", "Notes"]
    _header(ws, row, headers)
    header_row = row
    row += 1

    thin = Side(style="thin", color="D5D4CF")
    for i, day in enumerate(cal["days"]):
        values = [
            day["day"], day["date"], day["weekday"], day["post_type"],
            day["theme"], day["prompt"], day["hook"], day["best_time"],
            plain(day["confidence"]), "", "", "", "",
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row, col, value)
            c.alignment = Alignment(wrap_text=col in (5, 6, 7, 10, 13),
                                    vertical="top")
            c.border = Border(bottom=thin)
            if i % 2 == 1 and col <= 9:
                c.fill = PatternFill("solid", fgColor=BAND_BG)
            if col == 4:
                c.fill = PatternFill("solid", fgColor=TYPE_FILL[day["post_type"]])
                c.font = BOLD
            if col == 9:
                c.font = Font(color=VERDICT_COLOR.get(day["confidence"], MUTED),
                              size=10)
            if col in (10, 11, 12, 13):
                c.fill = PatternFill("solid", fgColor=ACCENT_BG)
        ws.row_dimensions[row].height = 40
        row += 1

    last = row - 1
    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(yn)
    yn.add(f"L{header_row + 1}:L{last}")

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    _stamp(ws, row + 1)
    _widths(ws, {1: 5, 2: 11, 3: 5, 4: 11, 5: 24, 6: 46, 7: 40, 8: 9, 9: 16,
                 10: 34, 11: 10, 12: 8, 13: 30})


def _hook_library(wb: Workbook, winning_hooks: list[dict], repetition: dict) -> None:
    ws = _sheet(wb, "Hook library")
    row = _title(ws, "Openings that already worked here",
                 "Taken from posts that beat this account's average, so the "
                 "voice is already yours.")
    _header(ws, row, ["Opening", "Views", "vs average", "Theme"])
    row += 1
    for h in winning_hooks:
        ws.cell(row, 1, h["hook"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 2, h["views"]).number_format = "#,##0"
        ws.cell(row, 3, h["index"] / 100).number_format = "0%"
        ws.cell(row, 4, h.get("theme") or "")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.cell(row, 1, "A warning about repeating yourself").font = BOLD
    row += 1
    if repetition.get("templated_posts"):
        note = (f"{repetition['templated_posts']} of your captions "
                f"({repetition['templated_share']}%) already open with a phrasing "
                f"you have used at least three times. Reusing the shape of a hook "
                f"is good. Reusing the exact words stops being a hook.")
    else:
        note = ("Your openings are not templated — that is a genuine strength, and "
                "worth protecting. Reuse the shape of what works, not the words.")
    c = ws.cell(row, 1, note)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 46
    _stamp(ws, row + 2)
    _widths(ws, {1: 76, 2: 12, 3: 12, 4: 26})


def _mix_tab(wb: Workbook, cal: dict) -> None:
    ws = _sheet(wb, "Content mix")
    row = _title(ws, "The mix across 30 days", "Six of each. That is the point.")
    _header(ws, row, ["Post type", "Posts", "Share", "What it does"])
    row += 1
    for m in cal["mix"]:
        c = ws.cell(row, 1, m["name"])
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor=TYPE_FILL[m["name"]])
        ws.cell(row, 2, m["n"])
        ws.cell(row, 3, m["share"] / 100).number_format = "0%"
        ws.cell(row, 4, m["does"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 38
        row += 1

    row += 1
    ws.cell(row, 1, "Themes in rotation").font = BOLD
    row += 1
    for t in cal["themes_used"]:
        ws.cell(row, 1, t)
        row += 1
    _stamp(ws, row + 1)
    _widths(ws, {1: 22, 2: 9, 3: 9, 4: 62})


def _progress(wb: Workbook, cal: dict) -> None:
    ws = _sheet(wb, "Progress tracker")
    row = _title(ws, "How the month went",
                 "Fill in the calendar tab and these update themselves.")
    first, last = 5, 4 + len(cal["days"])
    rows = [
        ("Days planned", len(cal["days"])),
        ("Posted so far", f'=COUNTIF(\'30-day calendar\'!L{first}:L{last},"Y")'),
        ("Still to go", f'=B{5} - B{6}'),
        ("Consistency", f'=IF(B{5}=0,0,B{6}/B{5})'),
    ]
    _header(ws, row, ["Measure", "Value"])
    row += 1
    start_row = row
    for label, value in rows:
        ws.cell(row, 1, label).font = BOLD
        ws.cell(row, 2, value)
        row += 1
    ws.cell(start_row + 3, 2).number_format = "0%"

    row += 1
    c = ws.cell(row, 1, "Consistency before performance. Thirty posts that went out "
                        "beat fifteen perfect ones that did not, because you cannot "
                        "learn what works from posts you never made.")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 48
    _stamp(ws, row + 2)
    _widths(ws, {1: 30, 2: 18})


# ------------------------------------------------------------------------ entry

def write(cal: dict, *, handle: str, winning_hooks: list[dict],
          repetition: dict, out_path: Path) -> Path:
    attribution.verify()
    wb = Workbook()
    wb.remove(wb.active)

    _how_to_use(wb, cal, handle)
    _post_type_guide(wb)
    _calendar_tab(wb, cal)
    _hook_library(wb, winning_hooks, repetition)
    _mix_tab(wb, cal)
    _progress(wb, cal)

    # Post-condition: every sheet must carry the credit line.
    for ws in wb.worksheets:
        text = "\n".join(str(c.value) for r in ws.iter_rows() for c in r
                         if c.value is not None)
        attribution.assert_stamped(text, f"calendar workbook sheet '{ws.title}'")

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
